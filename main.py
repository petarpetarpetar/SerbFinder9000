from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl import Workbook


conNumber = 0
url = ""
falseContacts = 0


class HeroldBot:

    def __init__(self,x):
        self.prezime = x
        self.driver = webdriver.Chrome("C:/chromedriver.exe")

    def closeBrowser(self):
        self.driver.close()

    def search(self):
        global url
        print("starting the search process")
        driver = self.driver
        driver.get("https://www.herold.at/")
        print("entering the website")
        time.sleep(1.5)
        print("connected, clicking on Person tab.")

        person_tab = driver.find_element_by_xpath("//*[@id='wp-tab-link']")
        person_tab.click()
        time.sleep(1.5)
        surname_element = driver.find_element_by_xpath("//*[@id='searchterm-wp']")
        surname_element.clear()
        surname_element.send_keys(self.prezime+"\n")  #\n is a faster way to execute the search
        time.sleep(1.5)
        #TODO:
        #make a city entry before executing the search.
        print("finished the search function")

        short_url = driver.current_url
        url = short_url + "?page=1"
        print("new url is " + url)
        driver.get(url)


    def gatherContacts(self):
        global excel
        global sheet1

        print("starting to gather Contracts")
        driver = self.driver
        try:
            temp = str(driver.find_element_by_css_selector(".col-sm-8.d-none.d-sm-block").get_attribute('innerHTML'))
        except:
            print("failed to obtain innerHTML information about pages, returning from function")
            return

        finally:
            print("successfully obtained data")

        totalPages = ((temp.split("von <b>")[1]).split("<")[0])
        for i in range(1,int(totalPages)+1):

            PageNumber = url.split("=")[0]
            PageNumber += "=" + str(i)
            driver.get(PageNumber)

            result_items_elements = []
            try:
                result_items_elements = driver.find_elements_by_class_name("result-item")
            except:
                print("couldn't find results")
                continue

            finally:
                print("found results")

            print("number of result items found on page " + str(i) + " is " + str(len(result_items_elements)))
            #print(result_items_elements)

            for i in range(0,len(result_items_elements)):
                global conNumber

                try:
                    address = result_items_elements[i].find_element_by_class_name("address").text
                    name = str(result_items_elements[i].find_element_by_class_name("col-lg-17").get_attribute('innerHTML'))
                    name = (name.split("name\">")[2]).split("</sp")[0]
                    telephone = str(result_items_elements[i].find_element_by_class_name("dropdown-item").get_attribute('innerHTML'))
                    telephone = telephone.split("</span>")[1]
                except:
                    print("FALSE CONTACT! ERROR")
                    time.sleep(2)
                    global falseContacts
                    falseContacts += 1

                finally:
                    conNumber += 1
                    print("#############CONTACT FOUND############### contact no. " + str((conNumber)))
                    print(address)
                    print(name)
                    print(telephone)
                    sheet1.cell(row=conNumber, column=1).value = conNumber
                    sheet1.cell(row=conNumber, column=2).value = name
                    sheet1.cell(row=conNumber, column=3).value = address
                    sheet1.cell(row=conNumber, column=4).value = telephone


excel = Workbook()
sheet1 = excel.active
f = open("prezime.txt","r")
for x in f:
    key = ''.join([i for i in x if not i.isdigit()])
    bot = HeroldBot(key)
    bot.search()
    bot.gatherContacts()
    bot.closeBrowser()
    del bot

print("found total of " + str(conNumber) + " correct contacts, and found " + str(falseContacts) + " corrupted/incorrect contacts.")
excel.save("data.xls")